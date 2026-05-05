import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="Fleet Status Report", layout="wide")
st.title("📊 Fleet Status by Project")


def normalize_status(value):
    if pd.isna(value) or str(value).strip() == "":
        return "Unknown"

    s = str(value).strip().lower()

    if s in ["ok", "running", "healthy", "online"]:
        return "OK"
    if "pending release" in s:
        return "Pending Release"
    if "down" in s or "offline" in s:
        return "Down"
    if "issue" in s or "warning" in s:
        return "Running with issues"

    return "Other"


def clean_dataframe(df):
    for col in df.columns:
        df[col] = df[col].apply(
            lambda x: str(int(x))
            if isinstance(x, float) and x.is_integer()
            else x
        )
    return df


# -------- LOAD EXCEL WITH HYPERLINKS --------
def load_workbook_with_links(uploaded_file):
    file_bytes = BytesIO(uploaded_file.read())
    wb = load_workbook(file_bytes, data_only=True)

    project_rows = []
    project_details = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        data = list(ws.values)
        if not data or len(data) < 2:
            continue

        df = pd.DataFrame(data[1:], columns=[str(c).strip() for c in data[0]])
        df = df.loc[:, ~df.columns.str.contains("^Unnamed", case=False)]
        df = df.fillna("")
        df = clean_dataframe(df)

        # --- Extract hyperlinks ---
        for col_name in ["OTE", "Extra", "Tracking"]:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name)

                for row in range(len(df)):
                    cell = ws.cell(row=row + 2, column=col_idx + 1)
                    if cell.hyperlink:
                        link = cell.hyperlink.target
                        df.iloc[row, col_idx] = f'<a href="{link}" target="_blank">{link}</a>'

        # --- Status normalization ---
        status_col = next((c for c in df.columns if "status" in c.lower()), None)
        if not status_col:
            continue

        normalized = df[status_col].apply(normalize_status)
        counts = normalized.value_counts()

        project_rows.append({
            "Project": sheet_name,
            "Fleet Size": len(df),
            "OK": counts.get("OK", 0),
            "Running with issues": counts.get("Running with issues", 0),
            "Pending Release": counts.get("Pending Release", 0),
            "Down": counts.get("Down", 0),
        })

        # --- Style status (bigger + colored) ---
        def style_status(val):
            txt = str(val)
            s = txt.lower()

            if s in ["ok", "running", "healthy", "online"]:
                color = "#7DCEA0"
            elif "issue" in s or "warning" in s:
                color = "#F5B041"
            elif "pending release" in s:
                color = "#85C1E9"
            elif "down" in s or "offline" in s:
                color = "#E59898"
            else:
                color = "inherit"

            return f'<span style="color:{color}; font-weight:bold; font-size:18px;">{txt}</span>'

        df[status_col] = df[status_col].apply(style_status)

        project_details[sheet_name] = df

    summary_df = pd.DataFrame(project_rows).sort_values("Fleet Size", ascending=False)
    return summary_df, project_details


# -------- FILE UPLOAD --------
uploaded = st.file_uploader("Upload Excel", type=["xlsx", "xls"])

if uploaded:

    summary_df, project_details = load_workbook_with_links(uploaded)

    # -------- GRAPH --------
    fig = go.Figure()

    fig.add_bar(x=summary_df["Project"], y=summary_df["OK"], name="OK", marker_color="#7DCEA0")
    fig.add_bar(x=summary_df["Project"], y=summary_df["Running with issues"], name="Issues", marker_color="#F5B041")
    fig.add_bar(x=summary_df["Project"], y=summary_df["Pending Release"], name="Pending", marker_color="#85C1E9")
    fig.add_bar(x=summary_df["Project"], y=summary_df["Down"], name="Down", marker_color="#E59898")

    fig.update_layout(
        barmode="stack",
        xaxis_title="<b>Project</b>",
        yaxis_title="<b># Robots</b>",
        height=550,
        xaxis_tickangle=-45
    )

    fig.update_xaxes(tickfont=dict(size=12, family="Arial Black"))

    st.plotly_chart(fig, use_container_width=True)

    # -------- PROJECT SELECT --------
    st.subheader("Project Summary")

    search = st.text_input("Search project")

    projects = [p for p in summary_df["Project"] if search.lower() in p.lower()]

    if not projects:
        st.warning("No project found")
        st.stop()

    selected_project = st.selectbox("Select project", projects)

    detail_df = project_details[selected_project]

    row = summary_df[summary_df["Project"] == selected_project].iloc[0]

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Fleet Size", int(row["Fleet Size"]))
    c2.metric("OK", int(row["OK"]))
    c3.metric("Issues", int(row["Running with issues"]))
    c4.metric("Pending", int(row["Pending Release"]))
    c5.metric("Down", int(row["Down"]))

    # -------- TABLE (FIXED RENDER) --------
    html_table = detail_df.to_html(index=False, escape=False)

    st.markdown(
        """
        <style>
        .fleet-table table {
            width: 100%;
            border-collapse: collapse;
            font-size: 15px;
        }
        .fleet-table th {
            text-align: left;
            padding: 8px;
            border-bottom: 2px solid #555;
        }
        .fleet-table td {
            padding: 8px;
            border-bottom: 1px solid #333;
            vertical-align: top;
            white-space: pre-wrap;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    st.markdown(
        f"""
        <div class="fleet-table" style="overflow-x:auto; height:700px;">
            {html_table}
        </div>
        """,
        unsafe_allow_html=True
    )

else:
    st.info("Upload your Excel file to generate the fleet status dashboard.")
